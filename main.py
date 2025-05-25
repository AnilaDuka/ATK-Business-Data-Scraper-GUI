import time
import tkinter as tk
from tkinter import messagebox
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
import os

def scrape_data(business_number):
    options = webdriver.ChromeOptions()
    driver = webdriver.Chrome(options=options)

    try:
        # Open the search page
        driver.get("https://apps.atk-ks.org/BizPasiveApp/VatRegist/Index")

        # Wait for page to load
        wait = WebDriverWait(driver, 60)
        search_input = wait.until(EC.element_to_be_clickable((By.ID, "txtFin")))

        driver.execute_script("arguments[0].scrollIntoView(true);", search_input)
        time.sleep(1)
        driver.execute_script("arguments[0].value = arguments[1];", search_input, business_number)

        # Let user solve the CAPTCHA manually
        messagebox.showinfo("CAPTCHA Required", "Please solve the CAPTCHA in the browser, then click OK here to continue.")

        # After user clicks OK, click the "Kërko" button automatically:
        search_button = wait.until(EC.element_to_be_clickable((By.ID, "Search")))
        search_button.click()
        time.sleep(3)

        # Wait for the results to load
        try:
            wait.until(EC.presence_of_element_located((By.ID, "FiscalNo")))
        except Exception as wait_err:
            messagebox.showerror("Wait Error", f"Timed out waiting for result field:\n{str(wait_err)}")
            driver.quit()
            return

        # Extract the info
        def get_text_by_id(element_id):
            try:
                return driver.find_element(By.ID, element_id).get_attribute("value")
            except:
                return ""
            
        data = {
            "Nr. Fiskal": get_text_by_id("FiscalNo"),
            "Emri": get_text_by_id("TpName"),
            "Adresa": get_text_by_id("Address"),
            "Qyteti": get_text_by_id("CityName"),
            "Komuna": get_text_by_id("ParishName"),
            "Qendra Tatimore": get_text_by_id("TaxCentreName"),
            "Nr. TVSh": get_text_by_id("VatNo"),
            "Statusi në TVSh": get_text_by_id("VatTypeAl"),
            "Tjetër": get_text_by_id("TpStatus")
        }
        print("Extracted data:", data)

        # Save to Excel
        save_to_excel(data)

        messagebox.showinfo("Success", "Data scraped and saved to Excel!")

    except Exception as e:
        messagebox.showerror("Error", f"Something went wrong:\n{str(e)}")
    finally:
        driver.quit()

def save_to_excel(data):
    file_name = "business_data.xlsx"

    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.append(list(data.keys())) # headers
    else:
        from openpyxl import load_workbook
        wb = load_workbook(file_name)
        ws = wb.active
    
    ws.append(list(data.values()))
    wb.save(file_name)


def on_submit():
    business_number = entry.get()
    if not business_number:
        messagebox.showwarning("Input Required", "Please enter a business number")
        return
    
    scrape_data(business_number)


# Create GUI
root = tk.Tk()
root.title("ATK Business Data Scraper")

tk.Label(root, text="Enter Business Number:").pack(pady=5)
entry = tk.Entry(root, width=30)
entry.pack(pady=5)

tk.Button(root, text="Scrape", command=on_submit).pack(pady=10)

root.geometry("300x150")
root.mainloop()